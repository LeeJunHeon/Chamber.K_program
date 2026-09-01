# lib/recipe_io.py
"""레시피 표 파일 로더 (CSV / TSV / XLSX 공용).

공정 레시피(main.py)와 히터 레시피(controller/heater_recipe.py)가 같이 쓴다.
어떤 형식이든 결과는 동일하게 list[dict] — 헤더가 키, 셀 값은 모두 문자열이다.
호출부의 파싱 로직을 그대로 쓰기 위한 것이므로, 여기서는 값을 해석하지 않는다.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Dict, Iterable, List

EXCEL_SUFFIXES = ('.xlsx', '.xlsm')
TEXT_SUFFIXES = ('.csv', '.tsv')


def _cell_to_str(v) -> str:
    """엑셀 셀 값을 문자열로. 정수형 float은 소수점을 떼어낸다.

    openpyxl은 5를 5.0으로 돌려주는데, 그대로 두면 "5.0"이 되어
    기존 CSV 파서(정수 기대)가 헷갈린다.
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return repr(v)
    return str(v).strip()


def _load_excel(path: Path, preferred_sheets: Iterable[str]) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "엑셀 레시피를 읽으려면 openpyxl 이 필요합니다: pip install openpyxl") from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = None
        for name in preferred_sheets or ():
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            if not wb.sheetnames:
                raise RuntimeError(f"시트가 없습니다: {path.name}")
            ws = wb[wb.sheetnames[0]]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise RuntimeError(f"시트 '{ws.title}' 가 비어 있습니다: {path.name}")

        headers = [_cell_to_str(h).strip() for h in header_row]
        if not any(headers):
            raise RuntimeError(f"시트 '{ws.title}' 의 1행에 헤더가 없습니다: {path.name}")

        out: List[Dict[str, str]] = []
        for raw in rows_iter:
            vals = [_cell_to_str(v) for v in raw]
            if not any(v.strip() for v in vals):
                continue        # 전부 빈 행은 건너뛴다
            row: Dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue    # 헤더 없는 열은 버린다
                row[h] = vals[i] if i < len(vals) else ""
            out.append(row)
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _load_text(path: Path) -> List[Dict[str, str]]:
    delim = '\t' if path.suffix.lower() == '.tsv' else ','
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=delim)
        return list(reader)


def load_table(path, preferred_sheets: Iterable[str] = ()) -> List[Dict[str, str]]:
    """표 파일을 읽어 list[dict] 로 돌려준다.

    - .csv / .tsv : csv.DictReader (utf-8-sig)
    - .xlsx / .xlsm : openpyxl. preferred_sheets 순서로 시트를 찾고 없으면 첫 시트
    - 그 외 확장자는 텍스트(csv)로 간주한다
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"레시피 파일이 없습니다: {p}")

    if p.suffix.lower() in EXCEL_SUFFIXES:
        return _load_excel(p, preferred_sheets)
    return _load_text(p)
